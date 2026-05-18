#!/usr/bin/env python3
import os
import glob
import sys
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json

# =============================================================================
# KONFIGURASI
# =============================================================================
DOSE_LIMIT        = 0.5
outputs = input('Masukkan nama folder:')
OUTPUT_DIR        = outputs
PHSP_PATTERN      = "DNADamage_*.phsp"
HEADER_FILE       = OUTPUT_DIR + "/" + "DNADamage.header"   # <-- diperbaiki: ada di folder outputs/
XLSX_OUTPUT       = OUTPUT_DIR + "/" + "hasil_simulasi.xlsx"

# Threshold dose minimum agar suatu segmen run dianggap valid.
# Segmen yang dose akhirnya di bawah nilai ini akan dibuang.
# Nilai 0.99 berarti run hanya diterima jika dosis kumulatif mencapai ≥99%.
DOSE_VALID_THRESHOLD = 0.49
# =============================================================================


def parse_header(header_file):
    col_names = []
    if not os.path.exists(header_file):
        return col_names
    with open(header_file, "r") as f:
        for line in f:
            line = line.strip()
            if ":" in line and line.split(":")[0].strip().isdigit():
                name = line.split(":", 1)[1].strip()
                col_names.append(name)
    return col_names


def split_into_segments(all_rows, dose_col_idx=1):
    """
    Memisahkan baris-baris data menjadi segmen-segmen run.

    Sebuah segmen baru dimulai ketika nilai dose pada suatu baris
    lebih kecil dari baris sebelumnya (yaitu terjadi reset/restart run).

    Returns:
        List of list-of-rows, satu elemen per segmen.
    """
    if not all_rows:
        return []

    segments = []
    current_seg = [all_rows[0]]

    for i in range(1, len(all_rows)):
        prev_dose = all_rows[i - 1][dose_col_idx] if dose_col_idx < len(all_rows[i - 1]) else 0
        curr_dose = all_rows[i][dose_col_idx]     if dose_col_idx < len(all_rows[i])     else 0
        if curr_dose < prev_dose:
            # Dosis turun → run baru dimulai, simpan segmen saat ini
            segments.append(current_seg)
            current_seg = [all_rows[i]]
        else:
            current_seg.append(all_rows[i])

    segments.append(current_seg)
    return segments


def parse_phsp_filtered(filepath, dose_limit, dose_col_idx=1,
                         dose_valid_threshold=DOSE_VALID_THRESHOLD):
    """
    Baca file .phsp dan kembalikan hanya baris dari segmen run yang valid.

    Sebuah segmen dianggap valid jika dose terakhirnya ≥ dose_valid_threshold.
    Segmen yang terpotong (dose akhir terlalu kecil) dibuang dan dilaporkan
    ke stdout sebagai peringatan.

    Args:
        filepath             : path ke file .phsp
        dose_limit           : (tidak dipakai saat ini, placeholder)
        dose_col_idx         : indeks kolom dose kumulatif (0-based), default 1
        dose_valid_threshold : ambang batas dose minimum agar segmen dianggap valid

    Returns:
        Tuple (valid_rows, stats_dict)
        - valid_rows : list of list-of-float, hanya dari segmen valid
        - stats_dict : info jumlah segmen total, valid, dibuang
    """
    all_rows = []
    with open(filepath, "r") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            parts = line.split()
            try:
                vals = [float(p) for p in parts]
            except ValueError:
                continue
            all_rows.append(vals)

    segments = split_into_segments(all_rows, dose_col_idx=dose_col_idx)

    valid_rows    = []
    n_valid       = 0
    n_discarded   = 0
    discarded_info = []
    fallback_used  = False

    for seg_idx, seg in enumerate(segments):
        # Ambil nilai dose baris terakhir segmen ini
        final_dose = seg[-1][dose_col_idx] if dose_col_idx < len(seg[-1]) else 0.0

        if final_dose >= dose_valid_threshold:
            valid_rows.extend(seg)
            n_valid += 1
        else:
            n_discarded += 1
            discarded_info.append(
                f"segmen {seg_idx + 1} ({len(seg)} baris, dose akhir = {final_dose:.6f})"
            )

    # Jika tidak ada segmen yang memenuhi threshold (misalnya file hanya berisi
    # sisa run yang terpotong), gunakan segmen dengan dose akhir tertinggi
    # sebagai fallback daripada mengembalikan data kosong (yang menghasilkan nol).
    if n_valid == 0 and segments:
        best_seg_idx = max(
            range(len(segments)),
            key=lambda i: (
                segments[i][-1][dose_col_idx]
                if dose_col_idx < len(segments[i][-1]) else 0.0
            )
        )
        best_seg       = segments[best_seg_idx]
        best_dose      = best_seg[-1][dose_col_idx] if dose_col_idx < len(best_seg[-1]) else 0.0
        valid_rows     = best_seg
        fallback_used  = True
        discarded_info.append(
            f"[FALLBACK] Tidak ada segmen yang mencapai threshold {dose_valid_threshold}. "
            f"Menggunakan segmen {best_seg_idx + 1} "
            f"({len(best_seg)} baris, dose akhir = {best_dose:.6f}) sebagai data terbaik."
        )

    stats = {
        "total_segments"    : len(segments),
        "valid_segments"    : n_valid,
        "discarded_segments": n_discarded,
        "discarded_info"    : discarded_info,
        "fallback_used"     : fallback_used,
    }
    return valid_rows, stats


def col_sum(rows, idx):
    return sum(row[idx] for row in rows if idx < len(row))

def col_last(rows, idx):
    """Ambil nilai baris terakhir pada kolom idx (untuk kolom kumulatif seperti dose)."""
    for row in reversed(rows):
        if idx < len(row):
            return row[idx]
    return None


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------
HEADER_FILL      = PatternFill("solid", start_color="1F4E79")
HEADER_FONT      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
SUBHEAD_FILL     = PatternFill("solid", start_color="D6E4F0")
SUBHEAD_FONT     = Font(name="Arial", bold=True, size=10)
DATA_FONT        = Font(name="Arial", size=10)
MEAN_FILL        = PatternFill("solid", start_color="E2EFDA")
STD_FILL         = PatternFill("solid", start_color="FFF2CC")
PARAM_FILL       = PatternFill("solid", start_color="F2F2F2")
PARAM_LABEL_FONT = Font(name="Arial", bold=True, size=10)

THIN         = Side(style="thin")
MED          = Side(style="medium")
THIN_BORDER  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MED_BORDER   = Border(left=MED,  right=MED,  top=MED,  bottom=MED)


def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border    = THIN_BORDER


def style_data_row(ws, row, ncols, fill=None):
    for c in range(1, ncols + 1):
        cell        = ws.cell(row=row, column=c)
        cell.font   = DATA_FONT
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill


def autofit(ws, min_w=8, max_w=25):
    for col in ws.columns:
        length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in col
        )
        ws.column_dimensions[col[0].column_letter].width = min(max(length + 2, min_w), max_w)


# ---------------------------------------------------------------------------
# Sheet Summary
# ---------------------------------------------------------------------------
def write_summary_sheet(wb, all_run_data, col_names, n_runs):
    ws          = wb.create_sheet(title="Summary")
    ncols_data  = len(col_names)
    ncols_total = ncols_data + 1   # +1 untuk kolom "Run ke-"

    # --- Baris 1: header ---
    ws.cell(row=1, column=1).value = "Run ke-"
    for ci, name in enumerate(col_names, start=2):
        ws.cell(row=1, column=ci).value = name
    style_header_row(ws, 1, ncols_total)

    # --- Baris 2..N+1: data tiap run ---
    data_start_row = 2
    for rd in all_run_data:
        r = data_start_row + rd["run"] - 1
        ws.cell(row=r, column=1).value = rd["run"]
        for ci, s in enumerate(rd["sums"], start=2):
            ws.cell(row=r, column=ci).value = round(s, 6) if s is not None else None
        style_data_row(ws, r, ncols_total)

    data_end_row = data_start_row + n_runs - 1

    # --- Baris Mean & Std ---
    mean_row = data_end_row + 2
    std_row  = mean_row + 1

    ws.cell(row=mean_row, column=1).value = "Mean"
    ws.cell(row=std_row,  column=1).value = "Std Dev"
    ws.cell(row=mean_row, column=1).font  = SUBHEAD_FONT
    ws.cell(row=std_row,  column=1).font  = SUBHEAD_FONT
    ws.cell(row=mean_row, column=1).fill  = MEAN_FILL
    ws.cell(row=std_row,  column=1).fill  = STD_FILL
    ws.cell(row=mean_row, column=1).border = THIN_BORDER
    ws.cell(row=std_row,  column=1).border = THIN_BORDER

    for ci in range(2, ncols_total + 1):
        col_letter = get_column_letter(ci)
        data_range = f"{col_letter}{data_start_row}:{col_letter}{data_end_row}"

        mean_cell = ws.cell(row=mean_row, column=ci)
        std_cell  = ws.cell(row=std_row,  column=ci)

        mean_cell.value          = f"=AVERAGE({data_range})"
        std_cell.value           = f"=STDEV({data_range})"
        mean_cell.font           = DATA_FONT
        std_cell.font            = DATA_FONT
        mean_cell.fill           = MEAN_FILL
        std_cell.fill            = STD_FILL
        mean_cell.border         = THIN_BORDER
        std_cell.border          = THIN_BORDER
        mean_cell.number_format  = "0.000000"
        std_cell.number_format   = "0.000000"

    # --- Tabel parameter tambahan ---
    # Kolom sheet = header_no (1-indexed) + 1 (karena kolom A = "Run ke-")
    def sc(header_1idx):
        return get_column_letter(header_1idx + 1)

    mr = mean_row  # baris Mean sebagai referensi formula

    params = [
        (
            "BD/SB",
            f"={sc(9)}{mr}/{sc(5)}{mr}",
            "BD/Gy/Gbp dibagi SB/Gy/Gbp"
        ),
        (
            "(Dir/Total)_SB",
            f"={sc(22)}{mr}/({sc(22)}{mr}+{sc(24)}{mr})",
            "SBs_Direct / (SBs_Direct + SBs_Indirect)"
        ),
        (
            "Dir/Total",
            (f"=({sc(22)}{mr}+{sc(29)}{mr})/"
             f"({sc(24)}{mr}+{sc(31)}{mr}+{sc(22)}{mr}+{sc(29)}{mr}"
             f"+{sc(23)}{mr}+{sc(30)}{mr})"),
            "(SBs_Direct+BDs_Direct) / (SBs_Indirect+BDs_Indirect+SBs_Direct+BDs_Direct+SBs_QuasiDirect+BDs_QuasiDirect)"
        ),
        (
            "(BD/SB)_Quasi",
            f"={sc(30)}{mr}/{sc(23)}{mr}",
            "BDs_QuasiDirect / SBs_QuasiDirect"
        ),
    ]

    param_start = std_row + 2
    title_row   = param_start

    ws.cell(row=title_row, column=1).value = "Parameter"
    ws.cell(row=title_row, column=2).value = "Nilai (dari Mean)"
    ws.cell(row=title_row, column=3).value = "Keterangan"
    for c in range(1, 4):
        cell           = ws.cell(row=title_row, column=c)
        cell.fill      = SUBHEAD_FILL
        cell.font      = SUBHEAD_FONT
        cell.border    = MED_BORDER
        cell.alignment = Alignment(horizontal="center")

    for i, (pname, formula, desc) in enumerate(params):
        r            = param_start + 1 + i
        label_cell   = ws.cell(row=r, column=1)
        formula_cell = ws.cell(row=r, column=2)
        desc_cell    = ws.cell(row=r, column=3)

        label_cell.value         = pname
        formula_cell.value       = formula
        desc_cell.value          = desc
        label_cell.font          = PARAM_LABEL_FONT
        formula_cell.font        = DATA_FONT
        desc_cell.font           = Font(name="Arial", italic=True, size=9)
        label_cell.fill          = PARAM_FILL
        formula_cell.fill        = PARAM_FILL
        label_cell.border        = THIN_BORDER
        formula_cell.border      = THIN_BORDER
        desc_cell.border         = THIN_BORDER
        formula_cell.number_format = "0.000000"

    # Garis batas tebal di sekeliling tabel parameter
    param_end = param_start + len(params)
    for r in range(title_row, param_end + 1):
        for c in range(1, 4):
            cell  = ws.cell(row=r, column=c)
            left  = MED if c == 1 else THIN
            right = MED if c == 3 else THIN
            top   = MED if r == title_row else THIN
            bot   = MED if r == param_end else THIN
            cell.border = Border(left=left, right=right, top=top, bottom=bot)

    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 18
    autofit(ws)
    ws.column_dimensions["C"].width = 65

    return ws

def load_runtime_data(output_dir):
    runtime_file = os.path.join(output_dir, "runtime_log.json")

    if not os.path.exists(runtime_file):
        return {}

    with open(runtime_file, "r") as f:
        raw = json.load(f)

    result = {}

    for entry in raw:
        result[entry["run"]] = entry["runtime"]

    return result

def write_runtime_sheet(wb, runtime_data):
    ws = wb.create_sheet(title="Runtime")

    headers = [
        "Run",
        "Parameter Reading Real (s)",
        "Initialization Real (s)",
        "Execution Real (s)",
        "Finalization Real (s)",
        "Total Real (s)",
    ]

    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c).value = h

    style_header_row(ws, 1, len(headers))

    for row_idx, run in enumerate(sorted(runtime_data.keys()), start=2):
        rt = runtime_data[run]

        ws.cell(row=row_idx, column=1).value = run
        ws.cell(row=row_idx, column=2).value = rt["Parameter Reading"]["Real"]
        ws.cell(row=row_idx, column=3).value = rt["Initialization"]["Real"]
        ws.cell(row=row_idx, column=4).value = rt["Execution"]["Real"]
        ws.cell(row=row_idx, column=5).value = rt["Finalization"]["Real"]
        ws.cell(row=row_idx, column=6).value = rt["Total"]["Real"]

        style_data_row(ws, row_idx, len(headers))

    autofit(ws)

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    files = sorted(glob.glob(os.path.join(OUTPUT_DIR, PHSP_PATTERN)))

    if not files:
        print(f"[ERROR] Tidak ada file .phsp di folder '{OUTPUT_DIR}'.")
        sys.exit(1)

    col_names = parse_header(HEADER_FILE)
    if not col_names:
        print(f"[WARNING] File header '{HEADER_FILE}' tidak ditemukan. "
              "Menggunakan nama kolom generik.")
        with open(files[0]) as f:
            for line in f:
                parts = line.strip().split()
                if parts:
                    col_names = [f"Col_{i+1}" for i in range(len(parts))]
                    break

    print(f"Ditemukan {len(files)} file .phsp")
    print(f"Jumlah kolom : {len(col_names)}")
    print(f"Threshold dose valid : ≥ {DOSE_VALID_THRESHOLD}")
    print()

    wb = Workbook()
    wb.remove(wb.active)  # hapus sheet default kosong
    runtime_data = load_runtime_data(OUTPUT_DIR)

    all_run_data        = []
    total_discarded_all = 0

    # Indeks kolom dose kumulatif (0-based); sesuai header: kolom ke-2 → indeks 1
    DOSE_COL_IDX = 1

    for run_idx, filepath in enumerate(files, start=1):
        fname = os.path.basename(filepath)
        rows, stats = parse_phsp_filtered(filepath, DOSE_LIMIT,
                                          dose_col_idx=DOSE_COL_IDX)

        n_total       = stats["total_segments"]
        n_valid       = stats["valid_segments"]
        n_discarded   = stats["discarded_segments"]
        fallback_used = stats["fallback_used"]
        total_discarded_all += n_discarded

        # Ringkasan per file
        if fallback_used:
            status = "FALLBACK: tidak ada segmen ≥ threshold, pakai segmen terbaik"
        elif n_discarded == 0:
            status = "OK"
        else:
            status = f"PERINGATAN: {n_discarded} segmen ekor dibuang"

        print(f"  [{run_idx:02d}] {fname}")
        print(f"        Segmen total={n_total}, valid={n_valid}, dibuang={n_discarded}  → {status}")
        for info in stats["discarded_info"]:
            print(f"              {info}")

        # Kolom dose bersifat kumulatif → ambil nilai baris terakhir segmen valid
        # Kolom lainnya dijumlahkan
        if rows:
            sums = [
                col_last(rows, ci) if ci == DOSE_COL_IDX else col_sum(rows, ci)
                for ci in range(len(col_names))
            ]
        else:
            # Seharusnya tidak pernah terjadi lagi setelah ada fallback,
            # tapi jaga-jaga jika file benar-benar kosong
            print(f"        [ERROR] File kosong atau tidak terbaca! Semua nilai = 0.")
            sums = [0.0] * len(col_names)

        all_run_data.append({
            "run" : run_idx,
            "sums": sums,
        })

    print()
    if total_discarded_all == 0:
        print("  ✓ Tidak ada segmen yang dibuang. Semua data bersih.")
    else:
        print(f"  ⚠ Total {total_discarded_all} segmen dibuang dari seluruh file "
              f"(threshold dose < {DOSE_VALID_THRESHOLD}).")

    print()
    print("  Membuat sheet Summary...")
    write_summary_sheet(wb, all_run_data, col_names, len(files))
    if runtime_data:
        print("  Membuat sheet Runtime...")
        write_runtime_sheet(wb, runtime_data)
    wb.save(XLSX_OUTPUT)
    print(f"\nFile tersimpan: {XLSX_OUTPUT}")


if __name__ == "__main__":
    main()
